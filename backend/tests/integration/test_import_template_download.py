import io
from zipfile import ZipFile

from fastapi.testclient import TestClient
from openpyxl import load_workbook


def test_download_record_import_template_zip(client: TestClient, auth_headers: dict[str, str]) -> None:
    response = client.get("/api/v1/imports/records/template/download", headers=auth_headers)

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/zip"

    archive = ZipFile(io.BytesIO(response.content))
    names = set(archive.namelist())
    assert names == {
        "records-template.json",
        "records-template.xlsx",
        "records-template.txt",
        "records-template.md",
    }

    json_template = archive.read("records-template.json").decode("utf-8")
    xlsx_template = archive.read("records-template.xlsx")
    txt_template = archive.read("records-template.txt").decode("utf-8")
    markdown_template = archive.read("records-template.md").decode("utf-8")
    workbook = load_workbook(io.BytesIO(xlsx_template), read_only=True, data_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))

    assert '"title": "Sample record 1"' in json_template
    assert rows[0] == ("title", "content")
    assert rows[1] == ("Sample record 1", "Completed the main task today and stayed steady overall.")
    assert "20250322" in txt_template
    assert "# Sample record 1" in markdown_template


def test_downloaded_txt_template_can_be_imported(client: TestClient, auth_headers: dict[str, str]) -> None:
    template_response = client.get("/api/v1/imports/records/template/download", headers=auth_headers)
    assert template_response.status_code == 200

    archive = ZipFile(io.BytesIO(template_response.content))
    txt_template = archive.read("records-template.txt")

    import_response = client.post(
        "/api/v1/imports/records",
        data={"source_type": "txt"},
        files={"file": ("records-template.txt", txt_template, "text/plain")},
        headers=auth_headers,
    )

    assert import_response.status_code == 201

    imports_response = client.get("/api/v1/imports/records", headers=auth_headers)
    task = imports_response.json()[0]
    assert task["status"] == "success"
    assert task["success_count"] == 2

    records_response = client.get("/api/v1/records", headers=auth_headers)
    titles = {item["title"] for item in records_response.json()}
    assert "20250322导入记录" in titles
    assert "250305导入记录" in titles
