from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.core.config import settings


DETAIL_TITLE = "请输入模板标题"
DETAIL_CONTENT = "请输入模板内容"


def test_template_default_is_unique_per_user(client: TestClient, auth_headers: dict[str, str]) -> None:
    first_response = client.post(
        "/api/v1/templates",
        json={"title": "Default A", "content": "A", "is_default": True},
        headers=auth_headers,
    )
    second_response = client.post(
        "/api/v1/templates",
        json={"title": "Default B", "content": "B", "is_default": True},
        headers=auth_headers,
    )

    assert first_response.status_code == 201
    assert second_response.status_code == 201

    list_response = client.get("/api/v1/templates", headers=auth_headers)
    assert list_response.status_code == 200
    templates = list_response.json()
    defaults = [item for item in templates if item["is_default"]]
    assert len(defaults) == 1
    assert defaults[0]["title"] == "Default B"


def test_template_validation_rejects_blank_title_and_content(client: TestClient, auth_headers: dict[str, str]) -> None:
    blank_title = client.post(
        "/api/v1/templates",
        json={"title": "   ", "content": "content", "is_default": False},
        headers=auth_headers,
    )
    blank_content = client.post(
        "/api/v1/templates",
        json={"title": "title", "content": "   ", "is_default": False},
        headers=auth_headers,
    )

    assert blank_title.status_code == 400
    assert blank_title.json()["detail"] == DETAIL_TITLE
    assert blank_content.status_code == 400
    assert blank_content.json()["detail"] == DETAIL_CONTENT


def test_template_import_json_file(client: TestClient, auth_headers: dict[str, str]) -> None:
    response = client.post(
        "/api/v1/templates/import",
        files={
            "upload_file": (
                "templates.json",
                '[{"title":"模板一","content":"内容一","is_default":true},{"title":"模板二","content":"内容二"}]'.encode("utf-8"),
                "application/json",
            )
        },
        headers=auth_headers,
    )

    assert response.status_code == 201
    assert response.json() == {"success_count": 2, "failed_count": 0, "total_count": 2}

    list_response = client.get("/api/v1/templates", headers=auth_headers)
    templates = list_response.json()
    assert len(templates) == 2
    assert templates[0]["is_default"] is True


def test_template_export_json_file(client: TestClient, auth_headers: dict[str, str]) -> None:
    create_response = client.post(
        "/api/v1/templates",
        json={"title": "导出模板", "content": "导出内容", "is_default": False},
        headers=auth_headers,
    )
    assert create_response.status_code == 201

    response = client.get("/api/v1/templates/export?format=json", headers=auth_headers)

    assert response.status_code == 200
    assert response.headers["content-disposition"] == 'attachment; filename="templates-export.json"'
    assert response.json()[0]["标题"] == "导出模板"
    assert "/" in response.json()[0]["创建时间"]


def test_template_import_xlsx_file(client: TestClient, auth_headers: dict[str, str]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["title", "content", "is_default"])
    sheet.append(["模板A", "内容A", True])
    sheet.append(["模板B", "内容B", False])
    output = BytesIO()
    workbook.save(output)

    response = client.post(
        "/api/v1/templates/import",
        files={
            "upload_file": (
                "templates.xlsx",
                output.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=auth_headers,
    )

    assert response.status_code == 201
    assert response.json() == {"success_count": 2, "failed_count": 0, "total_count": 2}


def test_template_export_xlsx_file(client: TestClient, auth_headers: dict[str, str]) -> None:
    create_response = client.post(
        "/api/v1/templates",
        json={"title": "模板导出", "content": "导出第一行\n导出第二行", "is_default": True},
        headers=auth_headers,
    )
    assert create_response.status_code == 201

    response = client.get("/api/v1/templates/export?format=xlsx", headers=auth_headers)

    assert response.status_code == 200
    assert response.headers["content-disposition"] == 'attachment; filename="templates-export.xlsx"'
    workbook = load_workbook(filename=BytesIO(response.content), read_only=True, data_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))
    assert rows[0] == ("标题", "内容", "默认模板", "创建时间", "更新时间")
    assert rows[1][0] == "模板导出"
    assert rows[1][1] == "导出第一行\n导出第二行"
    assert "/" in rows[1][3]


def test_template_import_rejects_oversized_upload(
    client: TestClient,
    auth_headers: dict[str, str],
    monkeypatch,
) -> None:
    monkeypatch.setattr(settings, "UPLOAD_MAX_FILE_SIZE_MB", 1)

    response = client.post(
        "/api/v1/templates/import",
        files={
            "upload_file": (
                "templates.json",
                b"x" * (1024 * 1024 + 1),
                "application/json",
            )
        },
        headers=auth_headers,
    )

    assert response.status_code == 413
    assert response.json()["detail"] == "Template import file exceeds the 1MB upload limit"
