import json
from datetime import UTC, date, datetime
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.analysis import Analysis
from app.models.audit_log import AuditLog
from app.models.backup_snapshot import BackupSnapshot
from app.models.export_task import ExportTask
from app.models.import_task import ImportTask
from app.models.record import Record
from app.models.template import Template


def test_create_record_import_persists_task_and_audit_log(
    client: TestClient,
    auth_headers: dict[str, str],
    db_session: Session,
) -> None:
    response = client.post(
        "/api/v1/imports/records",
        json={"source_type": "json", "file_name": "records.json"},
        headers=auth_headers,
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["source_type"] == "json"
    assert payload["file_name"] == "records.json"
    assert payload["status"] == "pending"
    assert payload["total_count"] == 0

    logs = db_session.query(AuditLog).all()
    assert len(logs) == 1
    assert logs[0].action == "create"
    assert logs[0].resource_type == "import_task"


def test_import_task_is_executed_in_background_and_generates_error_report(
    client: TestClient,
    auth_headers: dict[str, str],
) -> None:
    response = client.post(
        "/api/v1/imports/records",
        json={"source_type": "json", "file_name": "records.json"},
        headers=auth_headers,
    )

    assert response.status_code == 201

    imports_response = client.get("/api/v1/imports/records", headers=auth_headers)
    assert imports_response.status_code == 200

    task = imports_response.json()[0]
    assert task["status"] == "failed"
    assert task["failed_count"] == 1
    assert task["error_report_path"]

    report_response = client.get(f"/api/v1/imports/records/{task['id']}/report", headers=auth_headers)
    assert report_response.status_code == 200
    assert "did not provide uploaded file content" in report_response.text


def test_multipart_json_import_creates_records_and_marks_task_success(
    client: TestClient,
    auth_headers: dict[str, str],
) -> None:
    payload = json.dumps(
        [
            {"title": "Record A", "content": "Alpha"},
            {"title": "Record B", "content": "Beta"},
        ]
    ).encode("utf-8")

    response = client.post(
        "/api/v1/imports/records",
        data={"source_type": "json"},
        files={"file": ("records.json", payload, "application/json")},
        headers=auth_headers,
    )

    assert response.status_code == 201

    imports_response = client.get("/api/v1/imports/records", headers=auth_headers)
    records_response = client.get("/api/v1/records", headers=auth_headers)

    task = imports_response.json()[0]
    assert task["status"] == "success"
    assert task["success_count"] == 2
    assert task["failed_count"] == 0
    assert task["error_report_path"] is None

    records = records_response.json()
    assert len(records) == 2
    assert {item["title"] for item in records} == {"Record A", "Record B"}


def test_multipart_xlsx_import_creates_records(
    client: TestClient,
    auth_headers: dict[str, str],
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["title", "content"])
    sheet.append(["Morning", "Stayed calm"])
    sheet.append(["Evening", "Wrote notes"])
    output = BytesIO()
    workbook.save(output)
    xlsx_payload = output.getvalue()

    response = client.post(
        "/api/v1/imports/records",
        data={"source_type": "xlsx"},
        files={"file": ("records.xlsx", xlsx_payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=auth_headers,
    )

    assert response.status_code == 201

    imports_response = client.get("/api/v1/imports/records", headers=auth_headers)
    task = imports_response.json()[0]
    assert task["status"] == "success"
    assert task["success_count"] == 2


def test_record_import_rejects_oversized_upload(
    client: TestClient,
    auth_headers: dict[str, str],
    monkeypatch,
) -> None:
    monkeypatch.setattr(settings, "UPLOAD_MAX_FILE_SIZE_MB", 1)

    response = client.post(
        "/api/v1/imports/records",
        data={"source_type": "json"},
        files={"file": ("records.json", b"x" * (1024 * 1024 + 1), "application/json")},
        headers=auth_headers,
    )

    assert response.status_code == 413
    assert response.json()["detail"] == "Import file exceeds the 1MB upload limit"


def test_multipart_txt_import_splits_records_by_date_header(
    client: TestClient,
    auth_headers: dict[str, str],
) -> None:
    txt_payload = "\n".join(
        [
            "20250322",
            "第一条记录内容",
            "补充说明",
            "",
            "250305",
            "第二条记录内容",
        ]
    ).encode("utf-8")

    response = client.post(
        "/api/v1/imports/records",
        data={"source_type": "txt"},
        files={"file": ("journal.txt", txt_payload, "text/plain")},
        headers=auth_headers,
    )

    assert response.status_code == 201

    records_response = client.get("/api/v1/records", headers=auth_headers)
    records = records_response.json()
    assert len(records) == 2
    titles = {item["title"] for item in records}
    assert titles == {"20250322导入记录", "250305导入记录"}
    contents = {item["content"] for item in records}
    assert "第一条记录内容\n补充说明" in contents
    assert "第二条记录内容" in contents
    created_times = {item["created_at"][:16] for item in records}
    assert created_times == {"2025-03-22T22:00", "2025-03-05T22:00"}


def test_multipart_txt_import_supports_gbk_text(
    client: TestClient,
    auth_headers: dict[str, str],
) -> None:
    txt_payload = "\n".join(
        [
            "20250322",
            "第一条记录内容",
            "补充说明",
        ]
    ).encode("gb18030")

    response = client.post(
        "/api/v1/imports/records",
        data={"source_type": "txt"},
        files={"file": ("journal.txt", txt_payload, "text/plain")},
        headers=auth_headers,
    )

    assert response.status_code == 201

    imports_response = client.get("/api/v1/imports/records", headers=auth_headers)
    task = imports_response.json()[0]
    assert task["status"] == "success"
    assert task["success_count"] == 1

    records_response = client.get("/api/v1/records", headers=auth_headers)
    records = records_response.json()
    assert len(records) == 1
    assert records[0]["content"] == "第一条记录内容\n补充说明"


def test_create_export_and_backup_are_returned_by_jobs(
    client: TestClient,
    auth_headers: dict[str, str],
) -> None:
    export_response = client.post(
        "/api/v1/exports",
        json={"export_type": "records", "format": "json"},
        headers=auth_headers,
    )
    backup_response = client.post(
        "/api/v1/backups",
        json={"format": "zip"},
        headers=auth_headers,
    )

    assert export_response.status_code == 201
    assert backup_response.status_code == 201

    jobs_response = client.get("/api/v1/jobs", headers=auth_headers)

    assert jobs_response.status_code == 200
    jobs = jobs_response.json()
    assert len(jobs) == 2
    job_types = {item["type"] for item in jobs}
    assert "record_export" in job_types
    assert "backup_export" in job_types


def test_export_and_backup_tasks_are_executed_in_background(
    client: TestClient,
    auth_headers: dict[str, str],
    db_session: Session,
) -> None:
    record = Record(user_id=1, title="My Record", content="Today was steady.")
    template = Template(user_id=1, title="Default", content="Template body", is_default=True)
    analysis = Analysis(user_id=1, record_id=None, content="Analysis body", day_key=date(2026, 3, 16))
    db_session.add_all([record, template, analysis])
    db_session.commit()

    export_response = client.post(
        "/api/v1/exports",
        json={"export_type": "records", "format": "json"},
        headers=auth_headers,
    )
    backup_response = client.post(
        "/api/v1/backups",
        json={"format": "zip"},
        headers=auth_headers,
    )

    assert export_response.status_code == 201
    assert backup_response.status_code == 201

    exports_response = client.get("/api/v1/exports", headers=auth_headers)
    backups_response = client.get("/api/v1/backups", headers=auth_headers)

    assert exports_response.status_code == 200
    assert backups_response.status_code == 200

    export_task = exports_response.json()[0]
    backup_task = backups_response.json()[0]

    assert export_task["status"] == "success"
    assert export_task["file_path"]
    assert export_task["file_size"] > 0
    assert backup_task["status"] == "success"
    assert backup_task["storage_path"]
    assert backup_task["checksum"]

    export_path = export_task["file_path"]
    with open(export_path, "r", encoding="utf-8") as export_file:
        payload = json.load(export_file)
    assert payload[0]["标题"] == "My Record"
    assert payload[0]["创建时间"].startswith("20")

    backup_path = backup_task["storage_path"]
    with ZipFile(backup_path, "r") as archive:
        names = set(archive.namelist())
        assert names == {"records.xlsx", "templates.xlsx", "analyses.xlsx"}
        workbook = load_workbook(BytesIO(archive.read("templates.xlsx")), read_only=True, data_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
        assert rows[0] == ("ID", "标题", "内容", "默认模板", "创建时间", "更新时间")
        assert rows[1][1] == "Default"


def test_export_files_use_display_labels_and_render_multiline_content(
    client: TestClient,
    auth_headers: dict[str, str],
    db_session: Session,
) -> None:
    record = Record(user_id=1, title="多行记录", content="第一行\n第二行")
    analysis = Analysis(user_id=1, record_id=None, content="分析第一行\n分析第二行", day_key=date(2026, 3, 16))
    db_session.add_all([record, analysis])
    db_session.commit()

    records_export_response = client.post(
        "/api/v1/exports",
        json={"export_type": "records", "format": "txt"},
        headers=auth_headers,
    )
    analyses_export_response = client.post(
        "/api/v1/exports",
        json={"export_type": "analyses", "format": "markdown"},
        headers=auth_headers,
    )

    assert records_export_response.status_code == 201
    assert analyses_export_response.status_code == 201

    exports_response = client.get("/api/v1/exports", headers=auth_headers)
    exports = exports_response.json()
    records_task = next(item for item in exports if item["export_type"] == "records" and item["format"] == "txt")
    analyses_task = next(item for item in exports if item["export_type"] == "analyses" and item["format"] == "markdown")

    records_text = Path(records_task["file_path"]).read_text(encoding="utf-8")
    analyses_text = Path(analyses_task["file_path"]).read_text(encoding="utf-8")

    assert "标题：多行记录" in records_text
    assert "创建时间：" in records_text
    assert "第一行\n第二行" in records_text
    assert "\\n" not in records_text
    assert "### 内容" in analyses_text
    assert "分析第一行\n分析第二行" in analyses_text


def test_list_endpoints_are_user_scoped(
    client: TestClient,
    auth_headers: dict[str, str],
) -> None:
    client.post(
        "/api/v1/imports/records",
        json={"source_type": "xlsx", "file_name": "mine.xlsx"},
        headers=auth_headers,
    )

    client.post(
        "/api/v1/auth/register",
        json={"username": "other_user", "password": "other_password_123"},
    )
    other_login = client.post(
        "/api/v1/auth/login",
        json={"username": "other_user", "password": "other_password_123"},
    )
    other_headers = {"Authorization": f"Bearer {other_login.json()['access_token']}"}

    other_imports = client.get("/api/v1/imports/records", headers=other_headers)

    assert other_imports.status_code == 200
    assert other_imports.json() == []


def test_list_endpoints_keep_legacy_csv_tasks_readable(
    client: TestClient,
    auth_headers: dict[str, str],
    db_session: Session,
) -> None:
    import_task = ImportTask(user_id=1, source_type="csv", file_name="legacy.csv", status="success")
    export_task = ExportTask(user_id=1, export_type="records", format="csv", status="success")
    db_session.add_all([import_task, export_task])
    db_session.commit()

    imports_response = client.get("/api/v1/imports/records", headers=auth_headers)
    exports_response = client.get("/api/v1/exports", headers=auth_headers)

    assert imports_response.status_code == 200
    assert exports_response.status_code == 200
    assert imports_response.json()[0]["source_type"] == "csv"
    assert exports_response.json()[0]["format"] == "csv"


def test_get_job_returns_404_for_unknown_job(
    client: TestClient,
    auth_headers: dict[str, str],
) -> None:
    response = client.get("/api/v1/jobs/999999", headers=auth_headers)

    assert response.status_code == 404
    assert response.json()["detail"] == "Job not found"


def test_download_routes_return_files_when_paths_exist(
    client: TestClient,
    auth_headers: dict[str, str],
    db_session: Session,
    tmp_path,
) -> None:
    import_task = ImportTask(user_id=1, source_type="json", file_name="records.json", status="failed")
    export_task = ExportTask(user_id=1, export_type="records", format="json", status="success")
    backup_snapshot = BackupSnapshot(user_id=1, format="zip", status="success")
    db_session.add_all([import_task, export_task, backup_snapshot])
    db_session.commit()
    db_session.refresh(import_task)
    db_session.refresh(export_task)
    db_session.refresh(backup_snapshot)

    report_path = tmp_path / "import-report.txt"
    report_path.write_text("import report", encoding="utf-8")
    export_path = tmp_path / "records.json"
    export_path.write_text('{"ok": true}', encoding="utf-8")
    backup_path = tmp_path / "backup.zip"
    backup_path.write_bytes(b"PK\x03\x04mock")

    import_task.error_report_path = str(report_path)
    export_task.file_path = str(export_path)
    backup_snapshot.storage_path = str(backup_path)
    backup_snapshot.created_at = datetime(2026, 3, 25, 17, 7, 55, tzinfo=UTC)
    db_session.add_all([import_task, export_task, backup_snapshot])
    db_session.commit()

    report_response = client.get(f"/api/v1/imports/records/{import_task.id}/report", headers=auth_headers)
    export_response = client.get(f"/api/v1/exports/{export_task.id}/download", headers=auth_headers)
    backup_response = client.get(f"/api/v1/backups/{backup_snapshot.id}/download", headers=auth_headers)

    assert report_response.status_code == 200
    assert report_response.text == "import report"
    assert export_response.status_code == 200
    assert export_response.text == '{"ok": true}'
    assert backup_response.status_code == 200
    assert backup_response.content == b"PK\x03\x04mock"
    assert 'filename="integration_user-20260325170755.zip"' in backup_response.headers["content-disposition"]


def test_backup_download_exposes_content_disposition_header_for_cors(
    client: TestClient,
    auth_headers: dict[str, str],
    db_session: Session,
    tmp_path,
) -> None:
    backup_snapshot = BackupSnapshot(user_id=1, format="zip", status="success")
    db_session.add(backup_snapshot)
    db_session.commit()
    db_session.refresh(backup_snapshot)

    backup_path = tmp_path / "backup.zip"
    backup_path.write_bytes(b"PK\x03\x04mock")
    backup_snapshot.storage_path = str(backup_path)
    db_session.add(backup_snapshot)
    db_session.commit()

    response = client.get(
        f"/api/v1/backups/{backup_snapshot.id}/download",
        headers={
            **auth_headers,
            "Origin": "http://127.0.0.1:5173",
        },
    )

    assert response.status_code == 200
    assert response.headers["access-control-expose-headers"] == "Content-Disposition"


def test_download_routes_return_404_when_paths_are_missing(
    client: TestClient,
    auth_headers: dict[str, str],
) -> None:
    report_response = client.get("/api/v1/imports/records/999999/report", headers=auth_headers)
    export_response = client.get("/api/v1/exports/999999/download", headers=auth_headers)
    backup_response = client.get("/api/v1/backups/999999/download", headers=auth_headers)

    assert report_response.status_code == 404
    assert export_response.status_code == 404
    assert backup_response.status_code == 404


def test_backup_import_restores_records_templates_and_analyses(
    client: TestClient,
    auth_headers: dict[str, str],
) -> None:
    def build_sheet(headers: list[str], rows: list[list[object]]) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    records_xlsx = build_sheet(
        ["ID", "标题", "内容", "创建时间", "更新时间"],
        [[101, "恢复记录", "恢复内容", "2026/03/18 10:00:00", "2026/03/18 11:00:00"]],
    )
    templates_xlsx = build_sheet(
        ["ID", "标题", "内容", "默认模板", "创建时间", "更新时间"],
        [[201, "恢复模板", "模板内容", "是", "2026/03/17 09:00:00", "2026/03/17 09:30:00"]],
    )
    analyses_xlsx = build_sheet(
        ["ID", "关联记录ID", "内容", "分析日期", "创建时间"],
        [[301, 101, "恢复分析", "2026/03/18", "2026/03/18 12:00:00"]],
    )

    zip_buffer = BytesIO()
    with ZipFile(zip_buffer, "w") as archive:
        archive.writestr("records.xlsx", records_xlsx)
        archive.writestr("templates.xlsx", templates_xlsx)
        archive.writestr("analyses.xlsx", analyses_xlsx)

    response = client.post(
        "/api/v1/backups/import",
        files={"file": ("backup.zip", zip_buffer.getvalue(), "application/zip")},
        headers=auth_headers,
    )

    assert response.status_code == 201
    assert response.json() == {"records_imported": 1, "templates_imported": 1, "analyses_imported": 1}

    records_response = client.get("/api/v1/records", headers=auth_headers)
    templates_response = client.get("/api/v1/templates", headers=auth_headers)
    analyses_response = client.get("/api/v1/analyses", headers=auth_headers)

    records = records_response.json()
    templates = templates_response.json()
    analyses = analyses_response.json()

    assert records[0]["title"] == "恢复记录"
    assert records[0]["created_at"].startswith("2026-03-18T10:00:00")
    assert templates[0]["title"] == "恢复模板"
    assert templates[0]["is_default"] is True
    assert analyses[0]["content"] == "恢复分析"


def test_backup_import_skips_existing_data_by_id(
    client: TestClient,
    auth_headers: dict[str, str],
) -> None:
    def build_sheet(headers: list[str], rows: list[list[object]]) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    zip_buffer = BytesIO()
    with ZipFile(zip_buffer, "w") as archive:
        archive.writestr(
            "records.xlsx",
            build_sheet(["ID", "标题", "内容", "创建时间", "更新时间"], [[701, "重复记录", "内容", "2026/03/18 10:00:00", "2026/03/18 11:00:00"]]),
        )
        archive.writestr(
            "templates.xlsx",
            build_sheet(["ID", "标题", "内容", "默认模板", "创建时间", "更新时间"], [[801, "重复模板", "模板", "是", "2026/03/17 09:00:00", "2026/03/17 09:30:00"]]),
        )
        archive.writestr(
            "analyses.xlsx",
            build_sheet(["ID", "关联记录ID", "内容", "分析日期", "创建时间"], [[901, 701, "重复分析", "2026/03/18", "2026/03/18 12:00:00"]]),
        )

    first_response = client.post(
        "/api/v1/backups/import",
        files={"file": ("backup.zip", zip_buffer.getvalue(), "application/zip")},
        headers=auth_headers,
    )
    second_response = client.post(
        "/api/v1/backups/import",
        files={"file": ("backup.zip", zip_buffer.getvalue(), "application/zip")},
        headers=auth_headers,
    )

    assert first_response.status_code == 201
    assert second_response.status_code == 201
    assert second_response.json() == {"records_imported": 0, "templates_imported": 0, "analyses_imported": 0}

    records_response = client.get("/api/v1/records", headers=auth_headers)
    templates_response = client.get("/api/v1/templates", headers=auth_headers)
    analyses_response = client.get("/api/v1/analyses", headers=auth_headers)

    assert len(records_response.json()) == 1
    assert len(templates_response.json()) == 1
    assert len(analyses_response.json()) == 1


def test_backup_import_rejects_oversized_upload(
    client: TestClient,
    auth_headers: dict[str, str],
    monkeypatch,
) -> None:
    monkeypatch.setattr(settings, "UPLOAD_MAX_FILE_SIZE_MB", 1)

    response = client.post(
        "/api/v1/backups/import",
        files={"file": ("backup.zip", b"x" * (1024 * 1024 + 1), "application/zip")},
        headers=auth_headers,
    )

    assert response.status_code == 413
    assert response.json()["detail"] == "Backup file exceeds the 1MB upload limit"


def test_delete_task_endpoints_remove_owned_tasks(
    client: TestClient,
    auth_headers: dict[str, str],
    db_session: Session,
    tmp_path,
) -> None:
    import_task = ImportTask(user_id=1, source_type="json", file_name="records.json", status="failed")
    export_task = ExportTask(user_id=1, export_type="records", format="json", status="success")
    backup_snapshot = BackupSnapshot(user_id=1, format="zip", status="success")
    db_session.add_all([import_task, export_task, backup_snapshot])
    db_session.commit()
    db_session.refresh(import_task)
    db_session.refresh(export_task)
    db_session.refresh(backup_snapshot)

    report_path = tmp_path / "import-report.txt"
    report_path.write_text("import report", encoding="utf-8")
    export_path = tmp_path / "records.json"
    export_path.write_text('{"ok": true}', encoding="utf-8")
    backup_path = tmp_path / "backup.zip"
    backup_path.write_bytes(b"PK\x03\x04mock")

    import_task.error_report_path = str(report_path)
    export_task.file_path = str(export_path)
    backup_snapshot.storage_path = str(backup_path)
    db_session.add_all([import_task, export_task, backup_snapshot])
    db_session.commit()

    import_task_id = import_task.id
    export_task_id = export_task.id
    backup_snapshot_id = backup_snapshot.id

    delete_import = client.delete(f"/api/v1/imports/records/{import_task_id}", headers=auth_headers)
    delete_export = client.delete(f"/api/v1/exports/{export_task_id}", headers=auth_headers)
    delete_backup = client.delete(f"/api/v1/backups/{backup_snapshot_id}", headers=auth_headers)

    assert delete_import.status_code == 204
    assert delete_export.status_code == 204
    assert delete_backup.status_code == 204

    db_session.expire_all()
    assert db_session.get(ImportTask, import_task_id) is None
    assert db_session.get(ExportTask, export_task_id) is None
    assert db_session.get(BackupSnapshot, backup_snapshot_id) is None
    assert not report_path.exists()
    assert not export_path.exists()
    assert not backup_path.exists()


def test_delete_task_endpoints_are_user_scoped(
    client: TestClient,
    auth_headers: dict[str, str],
    db_session: Session,
) -> None:
    import_task = ImportTask(user_id=1, source_type="json", file_name="records.json", status="pending")
    export_task = ExportTask(user_id=1, export_type="records", format="json", status="pending")
    backup_snapshot = BackupSnapshot(user_id=1, format="zip", status="pending")
    db_session.add_all([import_task, export_task, backup_snapshot])
    db_session.commit()
    db_session.refresh(import_task)
    db_session.refresh(export_task)
    db_session.refresh(backup_snapshot)

    client.post("/api/v1/auth/register", json={"username": "other_delete_user", "password": "other_password_123"})
    other_login = client.post(
        "/api/v1/auth/login",
        json={"username": "other_delete_user", "password": "other_password_123"},
    )
    other_headers = {"Authorization": f"Bearer {other_login.json()['access_token']}"}

    delete_import = client.delete(f"/api/v1/imports/records/{import_task.id}", headers=other_headers)
    delete_export = client.delete(f"/api/v1/exports/{export_task.id}", headers=other_headers)
    delete_backup = client.delete(f"/api/v1/backups/{backup_snapshot.id}", headers=other_headers)

    assert delete_import.status_code == 404
    assert delete_export.status_code == 404
    assert delete_backup.status_code == 404
