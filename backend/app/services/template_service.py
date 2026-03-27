import io
import json
import re

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook

from app.core.config import settings
from app.core.upload_security import enforce_upload_size_limit
from app.models.template import Template
from app.repositories.template_repository import TemplateRepository
from app.schemas.template import TemplateCreate, TemplateUpdate
from app.services.export_formatting import (
    build_display_json_bytes,
    build_display_markdown,
    build_display_txt,
    build_display_xlsx_bytes,
)


class TemplateService:
    """处理模板的 CRUD 以及导入导出行为。"""
    def __init__(self, template_repository: TemplateRepository) -> None:
        self.template_repository = template_repository

    @staticmethod
    def validate_input(title: str, content: str) -> tuple[str, str]:
        """清洗用户输入，并拒绝空标题或空内容。"""
        clean_title = title.strip()
        clean_content = content.strip()

        if not clean_title:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="请输入模板标题")
        if not clean_content:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="请输入模板内容")

        return clean_title, clean_content

    def list_templates(self, user_id: int) -> list[Template]:
        return self.template_repository.list_by_user(user_id)

    def create_template(self, user_id: int, template_in: TemplateCreate) -> Template:
        """创建模板，并保证同一用户只有一个默认模板。"""
        title, content = self.validate_input(template_in.title, template_in.content)
        if template_in.is_default:
            self.template_repository.clear_default_for_user(user_id)
        return self.template_repository.create(user_id, title, content, template_in.is_default)

    def update_template(self, template_id: int, user_id: int, template_in: TemplateUpdate) -> Template:
        """在校验归属和内容后更新模板。"""
        template = self.template_repository.get_by_id_for_user(template_id, user_id)
        if template is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Template not found")

        title, content = self.validate_input(template_in.title, template_in.content)
        if template_in.is_default:
            self.template_repository.clear_default_for_user(user_id)
        return self.template_repository.update(template, title, content, template_in.is_default)

    def delete_template(self, template_id: int, user_id: int) -> None:
        template = self.template_repository.get_by_id_for_user(template_id, user_id)
        if template is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Template not found")
        self.template_repository.delete(template)

    async def import_templates(self, user_id: int, upload_file: UploadFile) -> dict[str, int]:
        """从一个上传文件中导入模板。"""
        file_name = upload_file.filename or "templates"
        file_bytes = enforce_upload_size_limit(
            await upload_file.read(),
            settings.upload_max_file_size_bytes,
            label="Template import file",
        )
        payload = self._parse_import_file(file_name, file_bytes)

        success_count = 0
        failed_count = 0
        for item in payload:
            if not isinstance(item, dict):
                failed_count += 1
                continue

            title = str(item.get("title", item.get("标题", "")))
            content = str(item.get("content", item.get("内容", "")))
            is_default = self._normalize_is_default(item)

            try:
                clean_title, clean_content = self.validate_input(title, content)
            except HTTPException:
                failed_count += 1
                continue

            if is_default:
                self.template_repository.clear_default_for_user(user_id)

            self.template_repository.create(user_id, clean_title, clean_content, is_default)
            success_count += 1

        if success_count == 0 and failed_count == 0:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="导入文件中没有可用的模板数据")

        return {
            "success_count": success_count,
            "failed_count": failed_count,
            "total_count": success_count + failed_count,
        }

    def export_templates(self, user_id: int, export_format: str) -> tuple[str, bytes]:
        """把模板序列化成指定导出格式。"""
        templates = self.template_repository.list_by_user(user_id)
        payload = [
            {
                "title": template.title,
                "content": template.content,
                "is_default": template.is_default,
                "created_at": template.created_at.isoformat(),
                "updated_at": template.updated_at.isoformat(),
            }
            for template in templates
        ]

        if export_format == "json":
            return "templates-export.json", build_display_json_bytes(payload)

        if export_format == "txt":
            return "templates-export.txt", build_display_txt("templates", payload)

        if export_format == "markdown":
            return "templates-export.md", build_display_markdown("templates", payload)

        if export_format == "xlsx":
            return "templates-export.xlsx", build_display_xlsx_bytes("templates", payload)

        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="不支持的模板导出格式")

    def _parse_import_file(self, file_name: str, file_bytes: bytes) -> list[dict[str, object]]:
        """根据上传文件后缀分发到对应解析器。"""
        normalized = file_name.lower()
        if normalized.endswith(".json"):
            return self._parse_json_templates(file_bytes)
        if normalized.endswith(".xlsx"):
            return self._parse_xlsx_templates(file_bytes)
        if normalized.endswith(".txt"):
            return self._parse_txt_templates(file_bytes)
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="模板导入暂只支持 json xlsx txt")

    def _parse_json_templates(self, file_bytes: bytes) -> list[dict[str, object]]:
        try:
            payload = json.loads(file_bytes.decode("utf-8-sig"))
        except (UnicodeDecodeError, json.JSONDecodeError) as exc:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="模板 JSON 文件格式无效") from exc

        items = payload.get("templates", []) if isinstance(payload, dict) else payload
        if not isinstance(items, list):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="模板 JSON 必须是数组或包含 templates 字段")
        return items

    def _parse_xlsx_templates(self, file_bytes: bytes) -> list[dict[str, object]]:
        try:
            workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception as exc:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="模板 XLSX 文件格式无效") from exc

        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        items: list[dict[str, object]] = []
        for values in rows[1:]:
            raw_item = {
                headers[position]: (values[position] if position < len(values) else None)
                for position in range(len(headers))
                if headers[position]
            }
            items.append(
                {
                    "title": raw_item.get("title", raw_item.get("标题")),
                    "content": raw_item.get("content", raw_item.get("内容")),
                    "is_default": raw_item.get("is_default", raw_item.get("默认模板")),
                }
            )
        return items

    def _parse_txt_templates(self, file_bytes: bytes) -> list[dict[str, object]]:
        text = self._decode_text_file(file_bytes)
        if text is None:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="模板 TXT 文件编码无效")

        normalized = text.replace("\r\n", "\n").strip()
        if normalized.startswith("【"):
            sections = re.split(r"(?m)^【[^】]+】\s*$", normalized)
            items: list[dict[str, object]] = []
            for body in sections[1:]:
                title_match = re.search(r"(?m)^标题：(.*)$", body)
                content_match = re.search(r"内容：\n(?P<content>.*?)(?:\n(?:创建时间：|更新时间：|默认模板：)|\Z)", body, re.S)
                title = title_match.group(1).strip() if title_match else ""
                content = content_match.group("content").strip() if content_match else ""
                if title and content:
                    items.append({"title": title, "content": content, "is_default": False})
            return items

        blocks = [block.strip() for block in text.split("\n\n") if block.strip()]
        items: list[dict[str, object]] = []
        for block in blocks:
            lines = [line.rstrip() for line in block.splitlines()]
            if len(lines) < 2:
                continue
            items.append(
                {
                    "title": lines[0].strip(),
                    "content": "\n".join(lines[1:]).strip(),
                    "is_default": False,
                }
            )
        return items

    @staticmethod
    def _decode_text_file(file_bytes: bytes) -> str | None:
        for encoding in ("utf-8-sig", "utf-8", "gb18030"):
            try:
                return file_bytes.decode(encoding)
            except UnicodeDecodeError:
                continue
        return None

    @staticmethod
    def _normalize_is_default(item: dict[str, object]) -> bool:
        raw_value = item.get("is_default", item.get("默认模板", False))
        if isinstance(raw_value, str):
            return raw_value.strip().lower() in {"true", "1", "yes", "y", "是"}
        return bool(raw_value)
