import json
from datetime import UTC, date, datetime
from io import BytesIO
from pathlib import Path
from zipfile import BadZipFile, ZipFile

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import update
from sqlalchemy.orm import Session

from app.models.analysis import Analysis
from app.models.backup_snapshot import BackupSnapshot
from app.models.record import Record
from app.models.template import Template
from app.repositories.backup_snapshot_repository import BackupSnapshotRepository
from app.core.config import settings
from app.core.upload_security import enforce_upload_size_limit
from app.schemas.backup import BackupCreate
from app.services.audit_service import AuditService


class BackupService:
    """负责备份快照的查询、下载、删除和数据恢复。"""

    def __init__(
        self,
        backup_snapshot_repository: BackupSnapshotRepository,
        db: Session,
        audit_service: AuditService | None = None,
    ) -> None:
        self.backup_snapshot_repository = backup_snapshot_repository
        self.db = db
        self.audit_service = audit_service

    def list_backups(self, user_id: int) -> list[BackupSnapshot]:
        """列出当前用户的全部备份快照。"""
        return self.backup_snapshot_repository.list_by_user(user_id)

    def create_backup(self, user_id: int, payload: BackupCreate) -> BackupSnapshot:
        """创建一个待执行的备份快照任务。"""
        snapshot = self.backup_snapshot_repository.create(user_id, payload.format)
        if self.audit_service is not None:
            self.audit_service.log(user_id, "create", "backup_snapshot", str(snapshot.id), f"format={payload.format}")
        return snapshot

    def get_backup_file_path(self, snapshot_id: int, user_id: int) -> Path:
        """获取备份文件路径，并校验归属和文件存在性。"""
        snapshot = self.backup_snapshot_repository.get_by_id_for_user(snapshot_id, user_id)
        if snapshot is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Backup snapshot not found")
        if not snapshot.storage_path:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Backup file not found")

        path = Path(snapshot.storage_path)
        if not path.exists() or not path.is_file():
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Backup file not found")
        return path

    def get_backup_download_name(self, snapshot_id: int, user_id: int, username: str) -> str:
        """生成备份下载文件名。"""
        snapshot = self.backup_snapshot_repository.get_by_id_for_user(snapshot_id, user_id)
        if snapshot is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Backup snapshot not found")

        timestamp = snapshot.created_at.strftime("%Y%m%d%H%M%S")
        safe_username = (username or "backup").strip() or "backup"
        return f"{safe_username}-{timestamp}.zip"

    def delete_backup(self, snapshot_id: int, user_id: int) -> None:
        """删除备份记录，并清理磁盘上的备份文件。"""
        snapshot = self.backup_snapshot_repository.get_by_id_for_user(snapshot_id, user_id)
        if snapshot is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Backup snapshot not found")

        file_path = Path(snapshot.storage_path) if snapshot.storage_path else None
        self.backup_snapshot_repository.delete(snapshot)
        if file_path and file_path.exists() and file_path.is_file():
            file_path.unlink(missing_ok=True)
        if self.audit_service is not None:
            self.audit_service.log(user_id, "delete", "backup_snapshot", str(snapshot_id), "delete backup snapshot")

    async def restore_backup(self, user_id: int, upload_file: UploadFile) -> dict[str, int]:
        """读取上传的备份压缩包，并把其中的数据恢复到当前用户账号下。"""
        file_bytes = enforce_upload_size_limit(
            await upload_file.read(),
            settings.upload_max_file_size_bytes,
            label="Backup file",
        )
        if not file_bytes:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Backup file is empty")

        try:
            archive = ZipFile(BytesIO(file_bytes))
        except BadZipFile as exc:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Backup file is not a valid ZIP archive") from exc

        with archive:
            payloads = self._load_backup_payloads(archive)

        template_id_map, templates_imported = self._restore_templates(user_id, payloads["templates"])
        record_id_map, records_imported = self._restore_records(user_id, payloads["records"], template_id_map)
        analyses_imported = self._restore_analyses(user_id, payloads["analyses"], record_id_map, template_id_map)
        self.db.commit()

        if self.audit_service is not None:
            self.audit_service.log(
                user_id,
                "restore",
                "backup_snapshot",
                upload_file.filename or "backup.zip",
                f"records={records_imported} templates={templates_imported} analyses={analyses_imported}",
            )

        return {
            "records_imported": records_imported,
            "templates_imported": templates_imported,
            "analyses_imported": analyses_imported,
        }

    def _load_backup_payloads(self, archive: ZipFile) -> dict[str, list[dict[str, object]]]:
        """兼容 JSON 和 XLSX 两种备份布局，统一转成字典列表。"""
        names = set(archive.namelist())
        if {"records.xlsx", "templates.xlsx", "analyses.xlsx"}.issubset(names):
            return {
                "records": self._load_xlsx_rows(archive.read("records.xlsx")),
                "templates": self._load_xlsx_rows(archive.read("templates.xlsx")),
                "analyses": self._load_xlsx_rows(archive.read("analyses.xlsx")),
            }
        if {"records.json", "templates.json", "analyses.json"}.issubset(names):
            return {
                "records": self._load_json_rows(archive.read("records.json")),
                "templates": self._load_json_rows(archive.read("templates.json")),
                "analyses": self._load_json_rows(archive.read("analyses.json")),
            }
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Backup archive does not contain a supported payload layout")

    @staticmethod
    def _load_json_rows(file_bytes: bytes) -> list[dict[str, object]]:
        """读取 JSON 数据，并确保根节点是对象数组。"""
        try:
            payload = json.loads(file_bytes.decode("utf-8"))
        except (UnicodeDecodeError, json.JSONDecodeError) as exc:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Backup JSON payload is invalid") from exc
        if not isinstance(payload, list):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Backup JSON payload must be a list")
        return [item for item in payload if isinstance(item, dict)]

    @staticmethod
    def _load_xlsx_rows(file_bytes: bytes) -> list[dict[str, object]]:
        """把 Excel 工作表按首行表头转换成字典列表。"""
        try:
            workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception as exc:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Backup XLSX payload is invalid") from exc

        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        result: list[dict[str, object]] = []
        for values in rows[1:]:
            result.append(
                {
                    headers[index]: (values[index] if index < len(values) else None)
                    for index in range(len(headers))
                    if headers[index]
                }
            )
        return result

    def _restore_records(
        self,
        user_id: int,
        rows: list[dict[str, object]],
        template_id_map: dict[int, int],
    ) -> tuple[dict[int, int], int]:
        """恢复记录数据，并建立旧记录 ID 到新记录 ID 的映射。"""
        record_id_map: dict[int, int] = {}
        imported = 0
        for row in rows:
            old_id = self._parse_int(self._get_raw(row, "id", "ID"))
            existing = self._get_record_by_source_id(user_id, old_id) if old_id is not None else None
            if existing is not None:
                record_id_map[old_id] = existing.id
                continue
            title = self._get_value(row, "title", "标题")
            content = self._get_value(row, "content", "内容")
            if not title or not content:
                continue
            created_at = self._parse_datetime(self._get_raw(row, "created_at", "创建时间"))
            updated_at = self._parse_datetime(self._get_raw(row, "updated_at", "更新时间"))
            old_template_id = self._parse_int(self._get_raw(row, "template_id", "关联模板ID"))
            record = Record(
                user_id=user_id,
                source_record_id=old_id,
                template_id=template_id_map.get(old_template_id) if old_template_id is not None else None,
                title=title,
                content=content,
                created_at=created_at,
                updated_at=updated_at,
            )
            self.db.add(record)
            self.db.flush()
            if old_id is not None:
                record_id_map[old_id] = record.id
            imported += 1
        return record_id_map, imported

    def _restore_templates(self, user_id: int, rows: list[dict[str, object]]) -> tuple[dict[int, int], int]:
        """恢复模板数据，同时重新计算默认模板归属。"""
        template_id_map: dict[int, int] = {}
        candidate_rows = [
            (index, row)
            for index, row in enumerate(rows)
            if not (
                (template_id := self._parse_int(self._get_raw(row, "id", "ID"))) is not None
                and self._get_template_by_source_id(user_id, template_id) is not None
            )
        ]
        default_indexes = [index for index, row in candidate_rows if self._parse_bool(self._get_raw(row, "is_default", "默认模板"))]
        last_default_index = default_indexes[-1] if default_indexes else None
        if last_default_index is not None:
            self.db.execute(update(Template).where(Template.user_id == user_id).values(is_default=False))

        imported = 0
        for index, row in enumerate(rows):
            template_id = self._parse_int(self._get_raw(row, "id", "ID"))
            if template_id is not None:
                existing = self._get_template_by_source_id(user_id, template_id)
                if existing is not None:
                    template_id_map[template_id] = existing.id
                    continue
            title = self._get_value(row, "title", "标题")
            content = self._get_value(row, "content", "内容")
            if not title or not content:
                continue
            template = Template(
                user_id=user_id,
                source_template_id=template_id,
                title=title,
                content=content,
                is_default=index == last_default_index,
                created_at=self._parse_datetime(self._get_raw(row, "created_at", "创建时间")),
                updated_at=self._parse_datetime(self._get_raw(row, "updated_at", "更新时间")),
            )
            self.db.add(template)
            self.db.flush()
            if template_id is not None:
                template_id_map[template_id] = template.id
            imported += 1
        return template_id_map, imported

    def _restore_analyses(
        self,
        user_id: int,
        rows: list[dict[str, object]],
        record_id_map: dict[int, int],
        template_id_map: dict[int, int],
    ) -> int:
        """恢复分析结果，并尽量把旧记录引用映射到新记录。"""
        imported = 0
        for row in rows:
            analysis_id = self._parse_int(self._get_raw(row, "id", "ID"))
            if analysis_id is not None and self._get_analysis_by_source_id(user_id, analysis_id) is not None:
                continue
            content = self._get_value(row, "content", "内容")
            if not content:
                continue
            day_key = self._parse_date(self._get_raw(row, "day_key", "分析日期"))
            created_at = self._parse_datetime(self._get_raw(row, "created_at", "创建时间")) or datetime.now(UTC)
            old_record_id = self._parse_int(self._get_raw(row, "record_id", "关联记录ID"))
            old_template_id = self._parse_int(self._get_raw(row, "template_id", "关联模板ID"))
            analysis = Analysis(
                user_id=user_id,
                source_analysis_id=analysis_id,
                record_id=record_id_map.get(old_record_id) if old_record_id is not None else None,
                template_id=template_id_map.get(old_template_id) if old_template_id is not None else None,
                content=content,
                day_key=day_key or created_at.date(),
                created_at=created_at,
            )
            self.db.add(analysis)
            imported += 1
        return imported

    @staticmethod
    def _get_raw(row: dict[str, object], *keys: str) -> object | None:
        """按多个候选字段名读取原始值，兼容中英文表头。"""
        for key in keys:
            if key in row:
                return row[key]
        return None

    @classmethod
    def _get_value(cls, row: dict[str, object], *keys: str) -> str:
        """读取字段并标准化为去空白字符串。"""
        raw = cls._get_raw(row, *keys)
        return str(raw).strip() if raw is not None else ""

    @staticmethod
    def _parse_bool(value: object | None) -> bool:
        """把导入文件中的布尔值文本转成 Python 布尔值。"""
        if isinstance(value, bool):
            return value
        if value is None:
            return False
        return str(value).strip().lower() in {"1", "true", "yes", "y", "是"}

    @staticmethod
    def _parse_int(value: object | None) -> int | None:
        """把可能为空的数字字段安全转成整数。"""
        if value is None or value == "":
            return None
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _parse_datetime(value: object | None) -> datetime | None:
        """兼容多种时间格式，失败时返回 None。"""
        if value is None or value == "":
            return None
        if isinstance(value, datetime):
            return value
        text = str(value).strip().replace("Z", "+00:00")
        try:
            return datetime.fromisoformat(text)
        except ValueError:
            pass
        for pattern in ("%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
            try:
                return datetime.strptime(str(value).strip(), pattern)
            except ValueError:
                continue
        return None

    @staticmethod
    def _parse_date(value: object | None) -> date | None:
        """兼容多种日期格式，失败时返回 None。"""
        if value is None or value == "":
            return None
        if isinstance(value, date) and not isinstance(value, datetime):
            return value
        if isinstance(value, datetime):
            return value.date()
        text = str(value).strip()
        for pattern in ("%Y/%m/%d", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, pattern).date()
            except ValueError:
                continue
        try:
            return date.fromisoformat(text)
        except ValueError:
            return None

    def _get_record_by_source_id(self, user_id: int, source_record_id: int) -> Record | None:
        """根据源记录 ID 查重，避免重复恢复。"""
        return (
            self.db.query(Record)
            .filter(Record.user_id == user_id, Record.source_record_id == source_record_id)
            .one_or_none()
        )

    def _get_template_by_source_id(self, user_id: int, source_template_id: int) -> Template | None:
        """根据源模板 ID 查重，避免重复恢复。"""
        return (
            self.db.query(Template)
            .filter(Template.user_id == user_id, Template.source_template_id == source_template_id)
            .one_or_none()
        )

    def _get_analysis_by_source_id(self, user_id: int, source_analysis_id: int) -> Analysis | None:
        """根据源分析 ID 查重，避免重复恢复。"""
        return (
            self.db.query(Analysis)
            .filter(Analysis.user_id == user_id, Analysis.source_analysis_id == source_analysis_id)
            .one_or_none()
        )
