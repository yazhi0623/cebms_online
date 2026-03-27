from __future__ import annotations

import hashlib
import io
import json
import re
from datetime import UTC, datetime, timedelta
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook
from sqlalchemy.orm import Session, sessionmaker

from app.models.analysis import Analysis
from app.models.record import Record
from app.models.template import Template
from app.repositories.analysis_repository import AnalysisRepository
from app.repositories.backup_snapshot_repository import BackupSnapshotRepository
from app.repositories.export_task_repository import ExportTaskRepository
from app.repositories.import_task_repository import ImportTaskRepository
from app.repositories.record_repository import RecordRepository
from app.repositories.template_repository import TemplateRepository
from app.services.export_formatting import (
    build_display_json_bytes,
    build_display_markdown,
    build_display_txt,
    build_display_xlsx_bytes,
)


def run_import_task(session_factory: sessionmaker[Session], task_id: int, file_bytes: bytes | None = None) -> None:
    """执行导入任务，把上传文件解析成记录数据并写入数据库。"""
    db = session_factory()
    task = None
    try:
        import_repo = ImportTaskRepository(db)
        record_repo = RecordRepository(db)
        template_repo = TemplateRepository(db)
        task = import_repo.get_by_id(task_id)
        if task is None:
            return

        import_repo.mark_running(task)
        # 导入文件的具体解析逻辑集中在这里分派，接口层只负责创建任务。
        imported_count, failed_count, report_lines = _process_import_file(
            task.source_type,
            task.file_name,
            file_bytes,
            record_repo,
            template_repo,
            task.user_id,
        )
        total_count = imported_count + failed_count
        report_path = None
        if report_lines:
            report_dir = _ensure_directory(Path(__file__).resolve().parents[2] / "generated" / "imports")
            report_path = report_dir / f"import-{task.id}-report.txt"
            report_path.write_text("\n".join(report_lines), encoding="utf-8")

        if failed_count == 0:
            import_repo.mark_success(task, total_count=total_count, success_count=imported_count, failed_count=0)
        else:
            import_repo.mark_failed(
                task,
                total_count=total_count,
                success_count=imported_count,
                failed_count=failed_count,
                error_report_path=str(report_path) if report_path is not None else None,
            )
    except Exception:
        if task is not None:
            ImportTaskRepository(db).mark_failed(task, total_count=0, success_count=0, failed_count=1)
        raise
    finally:
        db.close()


def run_export_task(session_factory: sessionmaker[Session], task_id: int) -> None:
    """执行导出任务，把数据库中的数据写成目标格式文件。"""
    db = session_factory()
    task = None
    try:
        export_repo = ExportTaskRepository(db)
        record_repo = RecordRepository(db)
        template_repo = TemplateRepository(db)
        analysis_repo = AnalysisRepository(db)
        task = export_repo.get_by_id(task_id)
        if task is None:
            return

        export_repo.mark_running(task)
        export_dir = _ensure_directory(Path(__file__).resolve().parents[2] / "generated" / "exports")
        payload = _get_export_payload(task.export_type, task.user_id, record_repo, template_repo, analysis_repo)
        file_path = export_dir / _build_export_filename(task_id, task.export_type, task.format)
        _write_export_file(file_path, task.export_type, task.format, payload)
        file_size = file_path.stat().st_size
        expires_at = datetime.now(UTC) + timedelta(days=7)
        export_repo.mark_success(task, str(file_path), file_size, expires_at)
    except Exception:
        if task is not None:
            ExportTaskRepository(db).mark_failed(task)
        raise
    finally:
        db.close()


def run_backup_task(session_factory: sessionmaker[Session], snapshot_id: int) -> None:
    """执行备份任务，把记录、模板、分析统一打成压缩包。"""
    db = session_factory()
    snapshot = None
    try:
        backup_repo = BackupSnapshotRepository(db)
        record_repo = RecordRepository(db)
        template_repo = TemplateRepository(db)
        analysis_repo = AnalysisRepository(db)
        snapshot = backup_repo.get_by_id(snapshot_id)
        if snapshot is None:
            return

        backup_repo.mark_running(snapshot)
        backup_dir = _ensure_directory(Path(__file__).resolve().parents[2] / "generated" / "backups")
        backup_path = backup_dir / f"backup-{snapshot.id}.zip"
        records = _serialize_records(record_repo.list_by_user(snapshot.user_id))
        templates = _serialize_templates(template_repo.list_by_user(snapshot.user_id))
        analyses = _serialize_analyses(analysis_repo.list_by_user(snapshot.user_id))

        with ZipFile(backup_path, "w", compression=ZIP_DEFLATED) as archive:
            archive.writestr("records.xlsx", build_display_xlsx_bytes("records", records))
            archive.writestr("templates.xlsx", build_display_xlsx_bytes("templates", templates))
            archive.writestr("analyses.xlsx", build_display_xlsx_bytes("analyses", analyses))

        checksum = hashlib.sha256(backup_path.read_bytes()).hexdigest()
        backup_repo.mark_success(snapshot, str(backup_path), checksum)
    except Exception:
        if snapshot is not None:
            BackupSnapshotRepository(db).mark_failed(snapshot)
        raise
    finally:
        db.close()


def _ensure_directory(path: Path) -> Path:
    """确保目标目录存在，并返回该目录对象。"""
    path.mkdir(parents=True, exist_ok=True)
    return path


def _process_import_file(
    source_type: str,
    file_name: str,
    file_bytes: bytes | None,
    record_repo: RecordRepository,
    template_repo: TemplateRepository,
    user_id: int,
) -> tuple[int, int, list[str]]:
    """根据导入源类型分派到对应解析器。"""
    if file_bytes is None:
        return (
            0,
            1,
            [
                f"file_name={file_name}",
                f"source_type={source_type}",
                "result=failed",
                "reason=Current request did not provide uploaded file content.",
                "next_step=Use multipart upload for /api/v1/imports/records to import real file data.",
            ],
        )

    if source_type == "json":
        return _import_json(file_name, file_bytes, record_repo, template_repo, user_id)
    if source_type == "xlsx":
        return _import_xlsx(file_name, file_bytes, record_repo, template_repo, user_id)
    if source_type == "txt":
        return _import_txt(file_name, file_bytes, record_repo, template_repo, user_id)
    if source_type == "markdown":
        return _import_markdown(file_name, file_bytes, record_repo, template_repo, user_id)
    return (
        0,
        1,
        [
            f"file_name={file_name}",
            "result=failed",
            f"reason=Unsupported source type: {source_type}",
        ],
    )


def _build_export_filename(task_id: int, export_type: str, export_format: str) -> str:
    """生成导出文件名，保持扩展名和导出格式一致。"""
    extension = "md" if export_format == "markdown" else export_format
    return f"{export_type}-export-{task_id}.{extension}"


def _get_export_payload(
    export_type: str,
    user_id: int,
    record_repo: RecordRepository,
    template_repo: TemplateRepository,
    analysis_repo: AnalysisRepository,
) -> list[dict[str, object]]:
    """读取指定资源类型的数据，并转换成可导出的字典列表。"""
    if export_type == "records":
        return _serialize_records(record_repo.list_by_user(user_id))
    if export_type == "templates":
        return _serialize_templates(template_repo.list_by_user(user_id))
    if export_type == "analyses":
        return _serialize_analyses(analysis_repo.list_by_user(user_id))
    raise ValueError(f"Unsupported export type: {export_type}")


def _write_export_file(path: Path, export_type: str, export_format: str, payload: list[dict[str, object]]) -> None:
    """按导出格式把载荷写到文件系统。"""
    if export_format == "json":
        path.write_bytes(build_display_json_bytes(payload))
        return

    if export_format == "txt":
        path.write_bytes(build_display_txt(export_type, payload))
        return

    if export_format == "markdown":
        path.write_bytes(build_display_markdown(export_type, payload))
        return

    if export_format == "xlsx":
        path.write_bytes(build_display_xlsx_bytes(export_type, payload))
        return

    raise ValueError(f"Unsupported export format: {export_format}")


def _import_json(
    file_name: str,
    file_bytes: bytes,
    record_repo: RecordRepository,
    template_repo: TemplateRepository,
    user_id: int,
) -> tuple[int, int, list[str]]:
    """解析 JSON 导入文件，支持数组和带 records 字段的对象。"""
    try:
        payload = json.loads(file_bytes.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError):
        return 0, 1, [f"file_name={file_name}", "result=failed", "reason=JSON file is not valid UTF-8 JSON."]

    items = payload.get("records", []) if isinstance(payload, dict) else payload
    if not isinstance(items, list):
        return 0, 1, [f"file_name={file_name}", "result=failed", "reason=JSON payload must be a list or an object with a records field."]

    success_count = 0
    failed_count = 0
    report_lines: list[str] = []
    for index, item in enumerate(items, start=1):
        if not isinstance(item, dict):
            failed_count += 1
            report_lines.append(f"row={index} reason=Item must be an object")
            continue
        title, content, created_at, updated_at, source_record_id, template_source_id = _extract_record_fields(item)
        template_id = _resolve_import_template_id(template_repo, user_id, template_source_id)
        if not title or not content:
            failed_count += 1
            report_lines.append(f"row={index} reason=Both title and content are required")
            continue
        if source_record_id is not None and record_repo.get_by_source_id_for_user(user_id, source_record_id) is not None:
            failed_count += 1
            report_lines.append(f"row={index} reason=Duplicate record ID {source_record_id} skipped")
            continue
        record_repo.create(
            user_id=user_id,
            title=title,
            content=content,
            template_id=template_id,
            created_at=created_at,
            updated_at=updated_at,
            source_record_id=source_record_id,
        )
        success_count += 1
    return success_count, failed_count, report_lines


def _import_xlsx(
    file_name: str,
    file_bytes: bytes,
    record_repo: RecordRepository,
    template_repo: TemplateRepository,
    user_id: int,
) -> tuple[int, int, list[str]]:
    """解析 Excel 导入文件，按首行表头映射字段。"""
    try:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        return 0, 1, [f"file_name={file_name}", "result=failed", "reason=XLSX file is invalid or unreadable."]

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return 0, 1, [f"file_name={file_name}", "reason=XLSX file is empty or missing headers"]

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    if not headers:
        return 0, 1, [f"file_name={file_name}", "reason=XLSX file is empty or missing headers"]

    success_count = 0
    failed_count = 0
    report_lines: list[str] = []
    for index, values in enumerate(rows[1:], start=2):
        row = {
            headers[position]: (values[position] if position < len(values) else None)
            for position in range(len(headers))
            if headers[position]
        }
        title, content, created_at, updated_at, source_record_id, template_source_id = _extract_record_fields(row)
        template_id = _resolve_import_template_id(template_repo, user_id, template_source_id)
        if not title or not content:
            failed_count += 1
            report_lines.append(f"row={index} reason=Both title and content columns are required")
            continue
        if source_record_id is not None and record_repo.get_by_source_id_for_user(user_id, source_record_id) is not None:
            failed_count += 1
            report_lines.append(f"row={index} reason=Duplicate record ID {source_record_id} skipped")
            continue
        record_repo.create(
            user_id=user_id,
            title=title,
            content=content,
            template_id=template_id,
            created_at=created_at,
            updated_at=updated_at,
            source_record_id=source_record_id,
        )
        success_count += 1

    if success_count == 0 and failed_count == 0:
        failed_count = 1
        report_lines.append(f"file_name={file_name} reason=XLSX file is empty or missing rows")
    return success_count, failed_count, report_lines


def _import_txt(
    file_name: str,
    file_bytes: bytes,
    record_repo: RecordRepository,
    template_repo: TemplateRepository,
    user_id: int,
) -> tuple[int, int, list[str]]:
    """解析 TXT 导入文件，兼容日期分段文本和导出后的行格式。"""
    text = _decode_text_file(file_bytes)
    if text is None:
        return 0, 1, [f"file_name={file_name}", "result=failed", "reason=TXT file is not valid UTF-8 or GBK text."]
    text = text.strip()
    if not text:
        return 0, 1, [f"file_name={file_name}", "reason=TXT file is empty"]

    exported_result = _import_txt_export_lines(text, record_repo, template_repo, user_id)
    if exported_result is not None:
        return exported_result

    pattern = re.compile(r"(?m)^(?P<date>\d{8}|\d{6})\s*$")
    matches = list(pattern.finditer(text))
    if not matches:
        return 0, 1, [f"file_name={file_name}", "reason=Each TXT record must start with a date line like 20250322 or 250305."]

    success_count = 0
    failed_count = 0
    report_lines: list[str] = []

    for index, match in enumerate(matches, start=1):
        date_text = match.group("date")
        content_start = match.end()
        content_end = matches[index].start() if index < len(matches) else len(text)
        content = text[content_start:content_end].strip()

        if not content:
            failed_count += 1
            report_lines.append(f"row={index} date={date_text} reason=Record content is empty")
            continue

        title = f"{date_text}导入记录"
        record_time = _parse_txt_record_datetime(date_text)
        record_repo.create(
            user_id=user_id,
            title=title[:100],
            content=content,
            created_at=record_time,
            updated_at=record_time,
        )
        success_count += 1

    return success_count, failed_count, report_lines


def _import_txt_export_lines(
    text: str,
    record_repo: RecordRepository,
    template_repo: TemplateRepository,
    user_id: int,
) -> tuple[int, int, list[str]] | None:
    """尝试把 TXT 导出内容重新导入；不匹配则返回 None。"""
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if not lines:
        return None

    section_result = _import_txt_display_sections(text, record_repo, template_repo, user_id)
    if section_result is not None:
        return section_result

    parsed_items: list[dict[str, object]] = []
    for line in lines:
        if not line.startswith("{"):
            return None
        try:
            item = json.loads(line)
        except json.JSONDecodeError:
            return None
        if not isinstance(item, dict):
            return None
        parsed_items.append(item)

    success_count = 0
    failed_count = 0
    report_lines: list[str] = []
    for index, item in enumerate(parsed_items, start=1):
        title, content, created_at, updated_at, source_record_id, template_source_id = _extract_record_fields(item)
        template_id = _resolve_import_template_id(template_repo, user_id, template_source_id)
        if not title or not content:
            failed_count += 1
            report_lines.append(f"row={index} reason=Both title and content are required")
            continue
        if source_record_id is not None and record_repo.get_by_source_id_for_user(user_id, source_record_id) is not None:
            failed_count += 1
            report_lines.append(f"row={index} reason=Duplicate record ID {source_record_id} skipped")
            continue
        record_repo.create(
            user_id=user_id,
            title=title,
            content=content,
            template_id=template_id,
            created_at=created_at,
            updated_at=updated_at,
            source_record_id=source_record_id,
        )
        success_count += 1

    return success_count, failed_count, report_lines


def _import_txt_display_sections(
    text: str,
    record_repo: RecordRepository,
    template_repo: TemplateRepository,
    user_id: int,
) -> tuple[int, int, list[str]] | None:
    """解析带展示标签的 TXT 导出内容。"""
    normalized = text.replace("\r\n", "\n").strip()
    if not normalized.startswith("【"):
        return None

    sections = re.split(r"(?m)^【[^\\n]+】\s*$", normalized)
    headers = re.findall(r"(?m)^【[^\\n]+】\s*$", normalized)
    if not headers:
        return None

    success_count = 0
    failed_count = 0
    report_lines: list[str] = []

    for index, body in enumerate(sections[1:], start=1):
        title_match = re.search(r"(?m)^标题：(.*)$", body)
        content_match = re.search(
            r"内容：\n(?P<content>.*?)(?:\n(?:创建时间：|更新时间：|默认模板：|分析日期：|关联记录ID：|关联模板ID：)|\Z)",
            body,
            re.S,
        )
        title = title_match.group(1).strip() if title_match else ""
        content = content_match.group("content").strip() if content_match else ""
        created_at = _parse_datetime_value(_match_labeled_value(body, "创建时间"))
        updated_at = _parse_datetime_value(_match_labeled_value(body, "更新时间"))
        source_record_id = _parse_optional_int(_match_labeled_value(body, "ID"))
        template_source_id = _parse_optional_int(_match_labeled_value(body, "关联模板ID"))
        template_id = _resolve_import_template_id(template_repo, user_id, template_source_id)
        if not title or not content:
            failed_count += 1
            report_lines.append(f"row={index} reason=Both title and content are required")
            continue
        if source_record_id is not None and record_repo.get_by_source_id_for_user(user_id, source_record_id) is not None:
            failed_count += 1
            report_lines.append(f"row={index} reason=Duplicate record ID {source_record_id} skipped")
            continue
        record_repo.create(
            user_id=user_id,
            title=title,
            content=content,
            template_id=template_id,
            created_at=created_at,
            updated_at=updated_at,
            source_record_id=source_record_id,
        )
        success_count += 1

    return success_count, failed_count, report_lines


def _import_markdown(
    file_name: str,
    file_bytes: bytes,
    record_repo: RecordRepository,
    template_repo: TemplateRepository,
    user_id: int,
) -> tuple[int, int, list[str]]:
    """解析 Markdown 导入文件，按标题拆分记录。"""
    text = _decode_text_file(file_bytes)
    if text is None:
        return 0, 1, [f"file_name={file_name}", "result=failed", "reason=Markdown file is not valid UTF-8 or GBK text."]

    sections = _parse_markdown_record_sections(text)
    if not sections:
        return 0, 1, [f"file_name={file_name}", "result=failed", "reason=Markdown file does not contain importable records."]

    success_count = 0
    failed_count = 0
    report_lines: list[str] = []
    fallback_title = Path(file_name).stem or "Markdown Import"

    for index, item in enumerate(sections, start=1):
        title = item["title"].strip() or f"{fallback_title}-{index}"
        content = item["content"].strip()
        created_at = _parse_datetime_value(item.get("created_at"))
        updated_at = _parse_datetime_value(item.get("updated_at"))
        source_record_id = _parse_optional_int(item.get("id"))
        template_source_id = _parse_optional_int(item.get("template_id"))
        template_id = _resolve_import_template_id(template_repo, user_id, template_source_id)
        if not content:
            failed_count += 1
            report_lines.append(f"row={index} reason=Record content is empty")
            continue
        if source_record_id is not None and record_repo.get_by_source_id_for_user(user_id, source_record_id) is not None:
            failed_count += 1
            report_lines.append(f"row={index} reason=Duplicate record ID {source_record_id} skipped")
            continue
        record_repo.create(
            user_id=user_id,
            title=title[:100],
            content=content,
            template_id=template_id,
            created_at=created_at,
            updated_at=updated_at,
            source_record_id=source_record_id,
        )
        success_count += 1

    return success_count, failed_count, report_lines


def _parse_markdown_record_sections(text: str) -> list[dict[str, str]]:
    """把 Markdown 文本拆成若干记录块。"""
    normalized = text.replace("\r\n", "\n").strip()
    if not normalized:
        return []

    exported_sections = _parse_exported_markdown_sections(normalized)
    if exported_sections:
        return exported_sections

    heading_pattern = re.compile(r"(?m)^(#{1,6})\s+(?P<title>.+?)\s*$")
    matches = list(heading_pattern.finditer(normalized))
    if not matches:
        plain_content = normalized.strip()
        return [{"title": "", "content": plain_content}] if plain_content else []

    sections: list[dict[str, str]] = []
    for index, match in enumerate(matches):
        title = match.group("title").strip()
        content_start = match.end()
        content_end = matches[index + 1].start() if index + 1 < len(matches) else len(normalized)
        content = normalized[content_start:content_end].strip()
        sections.append({"title": title, "content": content})
    return sections


def _parse_exported_markdown_sections(text: str) -> list[dict[str, str]]:
    """兼容解析系统导出的 Markdown 展示格式，也兼容旧的英文字段名。"""
    sections: list[dict[str, str]] = []
    current_fields: dict[str, str] = {}
    collecting_content = False
    content_lines: list[str] = []

    for line in text.split("\n"):
        stripped = line.strip()
        if not stripped:
            if collecting_content:
                content_lines.append("")
            continue

        if stripped.startswith("## "):
            if collecting_content:
                current_fields["内容"] = "\n".join(content_lines).strip()
                collecting_content = False
                content_lines = []
            if current_fields:
                parsed = _build_markdown_section(current_fields)
                if parsed["title"] or parsed["content"]:
                    sections.append(parsed)
            current_fields = {}
            continue

        if stripped.lower() in {"### 内容", "### content"}:
            collecting_content = True
            content_lines = []
            continue

        if collecting_content:
            if stripped.startswith("- **"):
                current_fields["内容"] = "\n".join(content_lines).strip()
                collecting_content = False
                content_lines = []
            else:
                content_lines.append(line.rstrip())
                continue

        field_match = re.match(r"^- \*\*(?P<key>[^*]+)\*\*[:：]\s*(?P<value>.*)$", stripped, re.IGNORECASE)
        if not field_match:
            return []

        key = field_match.group("key").strip().lower()
        value = field_match.group("value").strip()
        current_fields[key] = value

    if collecting_content:
        current_fields["内容"] = "\n".join(content_lines).strip()

    if current_fields:
        parsed = _build_markdown_section(current_fields)
        if parsed["title"] or parsed["content"]:
            sections.append(parsed)

    return [item for item in sections if item["title"] or item["content"]]


def _build_markdown_section(current_fields: dict[str, str]) -> dict[str, str]:
    """把 Markdown 解析阶段收集的字段标准化。"""
    return {
        "title": current_fields.get("title", current_fields.get("标题", "")).strip(),
        "content": current_fields.get("content", current_fields.get("内容", "")).strip(),
        "template_id": current_fields.get("template_id", current_fields.get("关联模板id", current_fields.get("关联模板ID", ""))).strip(),
        "created_at": current_fields.get("created_at", current_fields.get("创建时间", "")).strip(),
        "updated_at": current_fields.get("updated_at", current_fields.get("更新时间", "")).strip(),
    }


def _extract_record_fields(item: dict[str, object]) -> tuple[str, str, datetime | None, datetime | None, int | None, int | None]:
    """从导入行里提取记录核心字段，兼容中英文键名。"""
    title = str(item.get("title", item.get("标题", "")) or "").strip()
    content = str(item.get("content", item.get("内容", "")) or "").strip()
    created_at = _parse_datetime_value(item.get("created_at", item.get("创建时间")))
    updated_at = _parse_datetime_value(item.get("updated_at", item.get("更新时间")))
    source_record_id = _parse_optional_int(item.get("id", item.get("ID")))
    template_id = _parse_optional_int(item.get("template_id", item.get("关联模板ID")))
    return title, content, created_at, updated_at, source_record_id, template_id


def _parse_datetime_value(value: object) -> datetime | None:
    """兼容常见文本时间格式，失败时返回 None。"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    if not text:
        return None

    normalized = text.replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(normalized)
    except ValueError:
        pass

    for pattern in ("%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(text, pattern)
        except ValueError:
            continue
    return None


def _match_labeled_value(body: str, label: str) -> str | None:
    match = re.search(rf"(?m)^{re.escape(label)}[:：]\s*(.*)$", body)
    return match.group(1).strip() if match else None


def _parse_optional_int(value: object | None) -> int | None:
    """把可空字段安全转成整数。"""
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _resolve_import_template_id(template_repo: TemplateRepository, user_id: int, source_template_id: int | None) -> int | None:
    """把导入文件中的模板引用解析成本地模板 ID；若本地不存在则降级为空。"""
    if source_template_id is None:
        return None

    template = template_repo.get_by_source_id_for_user(user_id, source_template_id)
    if template is not None:
        return template.id

    template = template_repo.get_by_id_for_user(source_template_id, user_id)
    return template.id if template is not None else None


def _parse_txt_record_datetime(date_text: str) -> datetime:
    """把 TXT 导入中的日期行转成固定时间的 datetime。"""
    if len(date_text) == 8:
        parsed = datetime.strptime(date_text, "%Y%m%d")
    else:
        parsed = datetime.strptime(date_text, "%y%m%d")
    return parsed.replace(hour=22, minute=0, second=0, microsecond=0)


def _decode_text_file(file_bytes: bytes) -> str | None:
    """按常见中文文本编码依次尝试解码。"""
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return None


def _serialize_records(records: list[Record]) -> list[dict[str, object]]:
    """把记录模型序列化成可导出的字典。"""
    return [
        {
            "id": item.source_record_id or item.id,
            "template_id": item.template.source_template_id if item.template and item.template.source_template_id is not None else item.template_id,
            "title": item.title,
            "content": item.content,
            "created_at": item.created_at.isoformat(),
            "updated_at": item.updated_at.isoformat(),
        }
        for item in records
    ]


def _serialize_templates(templates: list[Template]) -> list[dict[str, object]]:
    """把模板模型序列化成可导出的字典。"""
    return [
        {
            "id": item.source_template_id or item.id,
            "title": item.title,
            "content": item.content,
            "is_default": item.is_default,
            "created_at": item.created_at.isoformat(),
            "updated_at": item.updated_at.isoformat(),
        }
        for item in templates
    ]


def _serialize_analyses(analyses: list[Analysis]) -> list[dict[str, object]]:
    """把分析模型序列化成可导出的字典。"""
    return [
        {
            "id": item.source_analysis_id or item.id,
            "record_id": item.record.source_record_id if item.record and item.record.source_record_id is not None else item.record_id,
            "template_id": item.template.source_template_id if item.template and item.template.source_template_id is not None else item.template_id,
            "content": item.content,
            "day_key": item.day_key.isoformat(),
            "created_at": item.created_at.isoformat(),
        }
        for item in analyses
    ]
